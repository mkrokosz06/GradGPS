import os

import openpyxl

# Repo root (two levels up from backend/scripts/) holds the Excel catalog
_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "PSU_Major_Requirements.xlsx")

wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True)
ws = wb['All Requirements']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
college_idx = headers.index('college')
prog_idx = headers.index('program_name')
degree_idx = headers.index('degree')

# Collect all UP programs (exclude Capital)
up_programs = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    prog = row[prog_idx]
    college = row[college_idx]
    degree = row[degree_idx]
    if prog and college != 'Capital':
        up_programs[prog] = (college, degree)

print(f"Total UP programs in Excel (excl Capital): {len(up_programs)}")

# Official UP bachelors from bulletin compiled by college
official_up_bachelors = {}

ag = [
    "Agribusiness Management, B.S.",
    "Agricultural and Biorenewable Systems Management, B.S.",
    "Agricultural and Extension Education, B.S.",
    "Agricultural Science, B.S.",
    "Animal Science, B.S.",
    "Community, Environment, and Development, B.S.",
    "Environmental Resource Management, B.S.",
    "Food Science, B.S.",
    "Forest Ecosystems, B.S.",
    "Immunology and Infectious Disease, B.S.",
    "Landscape Contracting, B.S.",
    "Pharmacology and Toxicology, B.S.",
    "Plant Sciences, B.S.",
    "Turfgrass Science, B.S.",
    "Veterinary and Biomedical Sciences, B.S.",
    "Wildlife and Fisheries Science, B.S.",
]
for p in ag: official_up_bachelors[p] = "Agricultural Sciences"

aa = [
    "Acting, B.F.A.",
    "Architecture, B.Arch.",
    "Architecture, B.S.",
    "Art Education, B.S.",
    "Art History, B.A.",
    "Art, B.A. (Arts and Architecture)",
    "Art, B.F.A.",
    "Digital Arts and Media Design, B.Des.",
    "Digital Multimedia Design, B.Des.",
    "Graphic Design, B.Des.",
    "Integrative Arts, B.A. (Arts and Architecture)",
    "Landscape Architecture, B.L.A.",
    "Music Education, B.M.E.",
    "Music Technology, B.M.",
    "Music, B.A.",
    "Music, B.M.",
    "Musical Arts, B.M.A.",
    "Musical Theatre, B.F.A.",
    "Professional Photography, B.Des.",
    "Theatre, B.A.",
    "Theatre, B.F.A.",
]
for p in aa: official_up_bachelors[p] = "Arts and Architecture"

bell = [
    "Advertising/Public Relations, B.A.",
    "Digital Journalism and Media, B.A.",
    "Film Production, B.A.",
    "Journalism, B.A.",
    "Media Studies, B.A.",
    "Strategic Communications, B.A.",
    "Telecommunications and Media Industries, B.A.",
]
for p in bell: official_up_bachelors[p] = "Bellisario Communications"

ems = [
    "Earth Science and Policy, B.S.",
    "Earth Sciences, B.S.",
    "Energy and Sustainability Policy, B.A.",
    "Energy and Sustainability Policy, B.S.",
    "Energy Business and Finance, B.S.",
    "Energy Engineering, B.S.",
    "Environmental Systems Engineering, B.S.",
    "Geobiology, B.S.",
    "Geography, B.A.",
    "Geography, B.S.",
    "Geosciences, B.A.",
    "Geosciences, B.S.",
    "Materials Science and Engineering, B.S.",
    "Meteorology and Atmospheric Science, B.S.",
    "Mining Engineering, B.S.",
    "Petroleum and Natural Gas Engineering, B.S.",
    "Sustainability, Society, and Environmental Geography, B.A.",
]
for p in ems: official_up_bachelors[p] = "Earth Mineral Sciences"

sci = [
    "Astronomy and Astrophysics, B.S.",
    "Biochemistry and Molecular Biology, B.S. (Science)",
    "Biology, B.S. (Science)",
    "Biotechnology, B.S.",
    "Chemistry, B.S. (Science)",
    "Data Sciences, B.S. (Science)",
    "Forensic Science, B.S.",
    "Integrative Science, B.S. (Science)",
    "Mathematics, B.A.",
    "Mathematics, B.S. (Science)",
    "Microbiology, B.S.",
    "Neurobiology, B.S.",
    "Physics, B.S. (Science)",
    "Planetary Science and Astronomy, B.S.",
    "Premedical-Medical, B.S.",
    "Premedicine, B.S.",
    "Science, B.S./Business Administration, M.B.A.",
    "Statistics, B.S.",
]
for p in sci: official_up_bachelors[p] = "Eberly Science"

edu = [
    "Education and Public Policy, B.S.",
    "Elementary and Early Childhood Education, B.S.",
    "Elementary and Kindergarten Education, B.S. (Education)",
    "Middle Level Education, B.S.",
    "Rehabilitation and Human Services, B.S. (Education)",
    "Secondary Education, B.S. (Education)",
    "Special Education, B.S.",
    "Workforce Education and Development, B.S.",
    "World Languages (K-12) Education, B.S.",
]
for p in edu: official_up_bachelors[p] = "Education"

eng = [
    "Aerospace Engineering, B.S.",
    "Architectural Engineering, B.A.E.",
    "Artificial Intelligence Engineering, B.S.",
    "Biological Engineering, B.S.",
    "Biomedical Engineering, B.S.",
    "Chemical Engineering, B.S.",
    "Civil Engineering, B.S. (Engineering)",
    "Computer Engineering, B.S. (Engineering)",
    "Computer Science, B.S. (Engineering)",
    "Data Sciences, B.S. (Engineering)",
    "Electrical Engineering Technology, B.S. (Engineering)",
    "Electrical Engineering, B.S. (Engineering)",
    "Electro-Mechanical Engineering Technology, B.S. (Engineering)",
    "Engineering Science, B.S.",
    "Engineering, B.S.",
    "Environmental Engineering, B.S.",
    "Industrial Engineering, B.S. (Engineering)",
    "Mechanical Engineering, B.S. (Engineering)",
    "Nuclear Engineering, B.S.",
    "Surveying Engineering, B.S.",
]
for p in eng: official_up_bachelors[p] = "Engineering"

hhd = [
    "Biobehavioral Health, B.S. (Health and Human Development)",
    "Communication Sciences and Disorders, B.S. (Health and Human Development)",
    "Health Policy and Administration, B.S. (Health and Human Development)",
    "Hospitality Management, B.S.",
    "Human Development and Family Studies, B.S. (Health and Human Development)",
    "Kinesiology, B.S.",
    "Nutritional Sciences, B.S.",
    "Recreation, Park, and Tourism Management, B.S. (Health and Human Development)",
    "Systems Neuroscience, B.S.",
]
for p in hhd: official_up_bachelors[p] = "Health and Human Development"

ist = [
    "Artificial Intelligence Methods and Applications, B.S. (Information Sciences and Technology)",
    "Cybersecurity Analytics and Operations, B.S. (Information Sciences and Technology)",
    "Data Sciences, B.S. (Information Sciences and Technology)",
    "Enterprise Technology Integration, B.S. (Information Sciences and Technology)",
    "Human-Centered Design and Development, B.S. (Information Sciences and Technology)",
    "Information Sciences and Technology, B.S.",
    "Information Technology Ethics and Compliance, B.S.",
    "Security and Risk Analysis, B.S. (Information Sciences and Technology)",
]
for p in ist: official_up_bachelors[p] = "Information Sciences and Technology"

inter = [
    "Bachelor of Philosophy Degree",
    "Business, B.S. (Intercollege)",
]
for p in inter: official_up_bachelors[p] = "Intercollege"

la = [
    "Administration of Justice, B.A.",
    "Administration of Justice, B.S.",
    "African American Studies, B.A.",
    "African Studies, B.A.",
    "African and African American Studies, B.A.",
    "African and African American Studies, B.S.",
    "Anthropological Science, B.S.",
    "Anthropology, B.A.",
    "Applied Linguistics, B.A.",
    "Asian Studies, B.A.",
    "Chinese, B.A.",
    "Classics and Ancient Mediterranean Studies, B.A.",
    "Communication Arts and Sciences, B.A. (Liberal Arts)",
    "Communication Arts and Sciences, B.S.",
    "Comparative Literature, B.A.",
    "Criminology, B.A.",
    "Criminology, B.S.",
    "Economics, B.A. (Liberal Arts)",
    "Economics, B.S.",
    "English, B.A. (Liberal Arts)",
    "French and Francophone Studies, B.A.",
    "French and Francophone Studies, B.S.",
    "German, B.A.",
    "German, B.S.",
    "Global and International Studies, B.A.",
    "Global and International Studies, B.S.",
    "History, B.A. (Liberal Arts)",
    "Integrated Social Sciences, B.S.",
    "International Politics, B.A.",
    "Italian, B.A.",
    "Italian, B.S.",
    "Japanese, B.A.",
    "Jewish Studies, B.A.",
    "Korean, B.A.",
    "Labor and Human Resources, B.A.",
    "Labor and Human Resources, B.S.",
    "Latin American Studies, B.A.",
    "Law and Society, B.A.",
    "Linguistics, B.A.",
    "Medieval Studies, B.A.",
    "Middle East Studies, B.A.",
    "Multidisciplinary Studies, B.A. (Liberal Arts)",
    "Organizational and Professional Communication, B.A.",
    "Organizational and Professional Communication, B.S.",
    "Organizational Leadership, B.A.",
    "Organizational Leadership, B.S.",
    "Philosophy, B.A.",
    "Philosophy, B.S.",
    "Political Science, B.A. (Liberal Arts)",
    "Political Science, B.S.",
    "Psychology, B.A. (Liberal Arts)",
    "Psychology, B.S. (Liberal Arts)",
    "Russian, B.A.",
    "Social Data Analytics, B.S.",
    "Sociology, B.A.",
    "Sociology, B.S. (Liberal Arts)",
    "Spanish, B.A.",
    "Spanish, B.S.",
    "Women's, Gender, and Sexuality Studies, B.A.",
    "Women's, Gender, and Sexuality Studies, B.S.",
]
for p in la: official_up_bachelors[p] = "Liberal Arts"

official_up_bachelors["Nursing, B.S.N."] = "Nursing"

smeal = [
    "Accounting, B.S. (Business)",
    "Actuarial Science, B.S.",
    "Business Analytics and Information Systems, B.S.",
    "Corporate Innovation and Entrepreneurship, B.S.",
    "Finance, B.S. (Business)",
    "Management, B.S. (Business)",
    "Marketing, B.S. (Business)",
    "Real Estate, B.S.",
    "Risk Management, B.S.",
    "Supply Chain and Information Systems, B.S.",
]
for p in smeal: official_up_bachelors[p] = "Smeal Business"

print(f"Official UP bachelor programs from bulletin: {len(official_up_bachelors)}")
print()

# Find what's in official list but NOT in our Excel (exact name match)
missing = []
for prog_name, college in official_up_bachelors.items():
    if prog_name not in up_programs:
        missing.append((prog_name, college))

print(f"MISSING from our Excel ({len(missing)} programs):")
for name, college in sorted(missing, key=lambda x: x[0]):
    print(f"  [{college}] {name}")

print()

# Find Excel bachelor degree programs NOT in official list
bachelor_degrees = {'B.S.', 'B.A.', 'B.F.A.', 'B.Arch.', 'B.Des.', 'B.M.', 'B.M.E.',
                    'B.M.A.', 'B.S.W.', 'B.S.N.', 'B.L.A.', 'B.A.E.', 'B.Phil.', 'N/A'}
excel_bachelors = {k: v for k, v in up_programs.items() if v[1] in bachelor_degrees}
not_in_official = []
for prog_name, (college, degree) in excel_bachelors.items():
    if prog_name not in official_up_bachelors:
        not_in_official.append((prog_name, college, degree))

print(f"In our Excel but NOT in official bulletin UP list ({len(not_in_official)} items):")
for name, college, degree in sorted(not_in_official, key=lambda x: x[0]):
    print(f"  [{college}] {degree} | {name}")

# Summary counts
print()
print("=== SUMMARY ===")
print(f"Official UP bachelor programs (bulletin): {len(official_up_bachelors)}")
print(f"Excel UP entries with bachelor degree type: {len(excel_bachelors)}")
print(f"Missing from Excel: {len(missing)}")
print(f"In Excel but not in official UP list: {len(not_in_official)}")
