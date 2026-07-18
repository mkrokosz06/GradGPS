"""
Penn State Undergraduate Major Requirements Scraper
Scrapes bulletins.psu.edu for all University Park undergraduate programs
and their course requirements. Outputs Excel + TXT files.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import re
import os

BASE_URL = "https://bulletins.psu.edu"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; academic-research-bot/1.0)"
}

# Global counter for unique pair group IDs across all programs
_pair_counter = [0]

def _next_pair_id():
    """Return a new globally-unique integer ID for an OR-alternative pair."""
    _pair_counter[0] += 1
    return _pair_counter[0]

def get_soup(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"  !! Failed: {url} — {e}")
                return None


# ── STEP 1: Get all undergraduate programs ──────────────────────────────────

def get_all_programs():
    print("Fetching program index from /programs/ ...")
    soup = get_soup(f"{BASE_URL}/programs/")
    programs = []
    if not soup:
        return programs

    seen = set()

    # Each program link is inside an <a> tag — get ONLY the link text,
    # not surrounding navigation. The link itself is the program name.
    for link in soup.find_all("a", href=True):
        href = link["href"]
        # Must match /undergraduate/colleges/[college]/[program]/
        if not re.match(r"^/undergraduate/colleges/[^/]+/[^/]+/?$", href):
            continue
        full_url = BASE_URL + href
        if full_url in seen:
            continue
        seen.add(full_url)

        # Clean name: only take the direct text of this <a>, strip trailing junk
        name = link.get_text(" ", strip=True)
        # Remove any text after common separators that indicate navigation bleed
        name = re.split(r"(?:Baccalaureate|Undergraduate|Graduate|Certificate|Minor|Penn State|University College)", name)[0].strip()
        name = re.sub(r"\s+", " ", name).strip(" ,")

        parts = href.strip("/").split("/")
        college_slug = parts[2] if len(parts) > 2 else "unknown"
        college = college_slug.replace("-", " ").title()

        programs.append({
            "name": name,
            "college": college,
            "college_slug": college_slug,
            "url": full_url
        })

    print(f"  Found {len(programs)} programs")
    return programs


# ── STEP 2: Scrape requirements for one program ──────────────────────────────

def scrape_program_requirements(program):
    soup = get_soup(program["url"])
    if not soup:
        return [], {}

    rows = []

    # ── Program metadata ──
    h1 = soup.find("h1")
    full_title = h1.get_text(strip=True) if h1 else program["name"]

    # Detect degree type
    degree = "N/A"
    for d in ["B.S.", "B.A.", "B.F.A.", "B.Arch.", "B.Des.", "B.Mus.", "B.Phil.", "B.Hum.", "B.Ed."]:
        if d in full_title:
            degree = d
            break
    if degree == "N/A":
        for d in ["Minor", "Certificate", "Associate"]:
            if d in full_title:
                degree = d
                break

    # Detect campus
    campus = "University Park"
    campus_indicators = ["Abington", "Altoona", "Behrend", "Berks", "Brandywine",
                         "DuBois", "Fayette", "Greater Allegheny", "Harrisburg",
                         "Hazleton", "Lehigh Valley", "Mont Alto", "New Kensington",
                         "Schuylkill", "Shenango", "Wilkes-Barre", "Worthington",
                         "York", "World Campus"]
    for c in campus_indicators:
        if c.lower() in full_title.lower() or c.lower() in program["url"].lower():
            campus = c
            break

    # ── Find the requirements tab content ──
    # PSU bulletin uses tab structure; look for the program requirements section
    req_content = None

    # Try common content containers
    for selector in [
        {"id": re.compile(r"requirementstab", re.I)},
        {"class": re.compile(r"(sc_page_content|program-requirements|tab-pane)", re.I)},
    ]:
        req_content = soup.find("div", selector)
        if req_content:
            break

    if not req_content:
        # Fall back to the whole page body
        req_content = soup.find("body") or soup

    # ── Parse requirement groups and courses ──
    current_group      = "General Requirements"
    current_group_type = "required"
    current_threshold  = None   # N for choose_credits or choose_courses

    def detect_group_type(text):
        """
        Classify a requirement group header into one of four types:
          required       - must complete every course listed
          choose_one     - take any one course from the list (either/or)
          choose_credits - pick courses until N credit hours are reached
          choose_courses - pick N courses from the list
        Also returns the numeric threshold (credits or course count) if present.
        """
        tl = text.lower()
        threshold = None

        # ── choose_credits: "choose N credits", "minimum N credits", "N credit hours" ──
        m = re.search(
            r"(?:choose|select|complete|minimum|at least)\s+(\d+)\s*(?:or more\s*)?credits?",
            tl
        )
        if m:
            return "choose_credits", int(m.group(1))

        # ── choose_courses: "choose N of", "select N courses", "complete N of" ──
        m = re.search(
            r"(?:choose|select|complete|take)\s+(\d+)\s+(?:of|course|from)",
            tl
        )
        if m:
            return "choose_courses", int(m.group(1))

        # ── choose_one: "one of the following", "or", option lists ──
        if any(p in tl for p in [
            "one of the following", "choose one", "select one",
            "complete one", "one of these", "either", " or "
        ]):
            return "choose_one", None

        # ── Gen-Ed ──
        if any(p in tl for p in ["general education", "gen ed", "gened", "university requirement",
                                   "knowledge domain", "integrative", "exploration", "foundation"]):
            return "gen_ed", None

        # ── Supporting / Related ──
        if any(p in tl for p in ["supporting", "related area", "elective"]):
            return "choose_credits", None   # usually a credit pool

        # ── Default: all required ──
        return "required", None

    # Walk through elements in order
    for el in req_content.find_all(["h2", "h3", "h4", "h5", "table"], recursive=True):
        tag = el.name

        # ── Section headers ──
        if tag in ["h2", "h3", "h4", "h5"]:
            text = el.get_text(strip=True)
            if len(text) < 4 or len(text) > 150:
                continue
            skip_words = ["admission", "suggested plan", "footnote", "note:", "sample plan",
                          "academic advising", "contact", "overview", "about", "career"]
            if any(s in text.lower() for s in skip_words):
                continue
            current_group      = text
            current_group_type, current_threshold = detect_group_type(text)

        # ── Also check paragraph/span text immediately before tables for pool instructions ──
        # e.g. "Select 3-4 credits from the following:"
        elif tag == "table":
            # Check preceding sibling text for "choose N credits" language
            prev = el.find_previous_sibling()
            if prev:
                prev_text = prev.get_text(" ", strip=True).lower()
                pg_type, pg_threshold = detect_group_type(prev_text)
                if pg_type in ("choose_credits", "choose_courses", "choose_one"):
                    current_group_type = pg_type
                    if pg_threshold:
                        current_threshold = pg_threshold

            # Track the last non-"or" row so we can group "or" alternatives with it
            last_course_idx = None   # index of last appended row
            current_pair_id = None   # pair_group_id shared by an OR-alternative set

            for tr in el.find_all("tr"):
                tds = tr.find_all(["td", "th"])
                if not tds:
                    continue

                cell_texts = [td.get_text(" ", strip=True) for td in tds]
                full_row   = " | ".join(cell_texts)
                first_cell = cell_texts[0].strip().lower()

                # ── Detect "or" rows: first cell is literally "or" or starts with "or " ──
                is_or_row = (
                    first_cell == "or"
                    or first_cell.startswith("or ")
                    or re.match(r"^or\s+[A-Z]{2,6}", cell_texts[0].strip())
                )

                # Must contain a course code somewhere in the row
                code_match = re.search(r"\b([A-Z]{2,6})\s{0,2}(\d{3}[A-Z]?)\b", full_row)
                if not code_match:
                    # Check for "select N credits" type rows inside the table
                    pool_match = re.search(
                        r"(?:select|choose|minimum)\s+(\d+)\s*(?:or more\s*)?credits?",
                        full_row, re.I
                    )
                    if pool_match:
                        current_group_type = "choose_credits"
                        current_threshold  = int(pool_match.group(1))
                    if not is_or_row:
                        current_pair_id = None   # break any open pair chain
                    continue

                course_code = f"{code_match.group(1)} {code_match.group(2)}"

                # Course title: usually second cell, cleaned
                title = cell_texts[1] if len(cell_texts) > 1 else cell_texts[0]
                title = re.sub(r"\b[A-Z]{2,6}\s{0,2}\d{3}[A-Z]?\b", "", title).strip(" –—/-or")
                title = re.sub(r"\s{2,}", " ", title).strip()
                if not title:
                    title = cell_texts[0]
                    title = re.sub(r"\b[A-Z]{2,6}\s{0,2}\d{3}[A-Z]?\b", "", title).strip()
                    title = re.sub(r"^\s*or\s*", "", title, flags=re.I).strip()

                # Credits
                credits = ""
                for ct in reversed(cell_texts[-2:]):
                    cm = re.search(r"^\s*(\d(?:\.\d+)?)\s*$", ct)
                    if cm:
                        val = float(cm.group(1))
                        if 0.5 <= val <= 9:
                            credits = cm.group(1)
                            break
                if not credits:
                    for ct in cell_texts:
                        cm = re.search(r"\b(\d(?:\.\d+)?)\s*cr", ct, re.I)
                        if cm:
                            credits = cm.group(1)
                            break

                # Min grade
                min_grade = ""
                row_lower = full_row.lower()
                if "grade of c" in row_lower or "c or better" in row_lower or "minimum c" in row_lower:
                    min_grade = "C"
                elif "grade of b" in row_lower or "b or better" in row_lower:
                    min_grade = "B"

                # ── pair_group_id: link OR-alternative courses together ──
                # Each "A / or B / or C" chain gets a single shared integer ID so
                # the audit engine knows exactly which courses are interchangeable.
                row_group_type = current_group_type
                this_pair_id   = None

                if is_or_row:
                    row_group_type = "choose_one"
                    if last_course_idx is not None:
                        rows[last_course_idx]["group_type"] = "choose_one"
                        if rows[last_course_idx]["pair_group_id"] is None:
                            # Start a brand-new pair: assign a fresh ID to both rows
                            current_pair_id = _next_pair_id()
                            rows[last_course_idx]["pair_group_id"] = current_pair_id
                        # else: extending an existing chain (A or B or C) — reuse the ID
                    this_pair_id = current_pair_id
                else:
                    # Non-OR row breaks any open pair chain
                    current_pair_id = None
                    this_pair_id    = None

                rows.append({
                    "program_name":      full_title,
                    "college":           program["college"].replace("-", " ").title(),
                    "degree":            degree,
                    "campus":            campus,
                    "requirement_group": current_group,
                    "group_type":        row_group_type,
                    "group_threshold":   current_threshold,
                    "course_code":       course_code,
                    "course_title":      title[:120],
                    "credits":           credits,
                    "min_grade":         min_grade,
                    "pair_group_id":     this_pair_id,
                    "url":               program["url"]
                })
                last_course_idx = len(rows) - 1

    meta = {
        "program_name": full_title,
        "college":      program["college"].replace("-", " ").title(),
        "degree":       degree,
        "campus":       campus,
        "url":          program["url"],
        "courses_found": len(rows)
    }

    return rows, meta


# ── STEP 3: Write styled Excel ───────────────────────────────────────────────

def write_excel(df, df_summary, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Requirements", index=False)
        df_summary.to_excel(writer, sheet_name="Programs Index", index=False)

        # One sheet per college (UP programs only to keep it manageable)
        df_up = df[df["campus"] == "University Park"]
        for college in sorted(df_up["college"].unique()):
            safe = re.sub(r"[^\w ]", "", college).strip()[:28]
            if safe:
                df_up[df_up["college"] == college].to_excel(
                    writer, sheet_name=safe, index=False
                )

    wb = openpyxl.load_workbook(path)

    hdr_fill  = PatternFill("solid", fgColor="1A3A6B")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    alt_fill  = PatternFill("solid", fgColor="EFF6FF")
    cell_font = Font(size=10, name="Segoe UI")
    wrap      = Alignment(wrap_text=True, vertical="top")
    center    = Alignment(horizontal="center", vertical="center")
    thin_bot  = Border(bottom=Side(style="thin", color="DBEAFE"))

    col_w = {
        "program_name":      40,
        "college":           22,
        "degree":             8,
        "campus":            16,
        "requirement_group": 30,
        "group_type":        14,
        "group_threshold":   14,
        "course_code":       13,
        "course_title":      46,
        "credits":            8,
        "min_grade":          9,
        "pair_group_id":     13,
        "url":                0,
        "name":              44,
        "courses_found":     15,
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Header row
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            ws.row_dimensions[1].height = 22

        # Column widths
        for ci, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            key = str(col_cells[0].value or "").lower().replace(" ", "_")
            w   = col_w.get(key, 18)
            col_ltr = get_column_letter(ci)
            if w == 0:
                ws.column_dimensions[col_ltr].hidden = True
            else:
                ws.column_dimensions[col_ltr].width = w

        # Data rows
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.font      = cell_font
                cell.alignment = wrap
                cell.border    = thin_bot
                if ri % 2 == 0:
                    cell.fill = alt_fill

    wb.save(path)


# ── STEP 4: Write TXT ────────────────────────────────────────────────────────

def write_txt(df, path):
    with open(path, "w", encoding="utf-8") as f:
        f.write("PENN STATE UNIVERSITY — UNDERGRADUATE MAJOR REQUIREMENTS\n")
        f.write("Source: bulletins.psu.edu\n")
        f.write("=" * 72 + "\n\n")

        current_prog  = None
        current_group = None

        sort_cols = ["campus", "college", "program_name", "requirement_group", "course_code"]
        for col in sort_cols:
            if col not in df.columns:
                sort_cols.remove(col)

        for _, row in df.sort_values(sort_cols).iterrows():
            prog_key = f"{row['program_name']} | {row['campus']}"

            if prog_key != current_prog:
                current_prog  = prog_key
                current_group = None
                f.write("\n" + "=" * 72 + "\n")
                f.write(f"  PROGRAM : {row['program_name']}\n")
                f.write(f"  DEGREE  : {row['degree']}\n")
                f.write(f"  COLLEGE : {row['college']}\n")
                f.write(f"  CAMPUS  : {row['campus']}\n")
                f.write("=" * 72 + "\n")

            if row["requirement_group"] != current_group:
                current_group = row["requirement_group"]
                f.write(f"\n  >> [{row['group_type']}]  {row['requirement_group']}\n")
                f.write(f"    {'─' * 55}\n")

            grade  = f"  <- min grade: {row['min_grade']}" if row["min_grade"] else ""
            cr     = f"{row['credits']} cr" if row["credits"] else "?"
            pair   = f"  [pair:{row['pair_group_id']}]" if row.get("pair_group_id") else ""
            f.write(f"    {row['course_code']:<14}  {cr:<6}  {row['course_title']}{pair}{grade}\n")

        f.write("\n\n" + "=" * 72 + "\n")
        total_progs = df["program_name"].nunique()
        total_rows  = len(df)
        f.write(f"  Total programs scraped : {total_progs}\n")
        f.write(f"  Total requirement rows : {total_rows}\n")
        f.write("=" * 72 + "\n")


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main(argv=None):
    import argparse

    parser = argparse.ArgumentParser(description="Scrape PSU bulletin → PSU_Major_Requirements.xlsx")
    parser.add_argument(
        "--program",
        action="append",
        default=[],
        help="Only scrape programs whose name contains this substring (repeatable). "
             "Example: --program 'Enterprise Technology'",
    )
    parser.add_argument(
        "--max",
        type=int,
        default=0,
        help="Stop after N University Park programs (0 = no limit)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List matching programs and exit without scraping requirements",
    )
    args = parser.parse_args(argv)

    # Repo root (two levels up from backend/scripts/) — load_catalog.py reads the
    # Excel output from there
    output_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))

    print("=" * 60)
    print("Penn State Major Requirements Scraper")
    print("=" * 60)

    programs = get_all_programs()
    if not programs:
        print("No programs found.")
        return

    if args.program:
        needles = [n.lower() for n in args.program]
        programs = [
            p for p in programs
            if any(n in p["name"].lower() for n in needles)
        ]
        print(f"  Filtered to {len(programs)} programs matching {args.program!r}")

    if args.dry_run:
        for p in programs:
            print(f"  - {p['name']}  ({p['college']})")
        print(f"Total: {len(programs)}")
        return

    all_rows    = []
    all_summary = []

    skipped = 0
    up_scraped = 0
    for i, prog in enumerate(programs, 1):
        print(f"[{i:>3}/{len(programs)}] {prog['name'][:65]}", end="", flush=True)
        rows, meta = scrape_program_requirements(prog)

        # Skip non-University Park programs
        if meta.get("campus", "University Park") != "University Park":
            print(f"  SKIP ({meta['campus']})")
            skipped += 1
            time.sleep(0.2)
            continue

        all_rows.extend(rows)
        all_summary.append(meta)
        up_scraped += 1
        print(f"  OK {len(rows)} courses")
        time.sleep(0.35)

        if args.max and up_scraped >= args.max:
            print(f"\nReached --max {args.max} University Park programs")
            break

    print(f"\nSkipped {skipped} non-University Park programs")

    print(f"\nTotal requirement rows collected: {len(all_rows)}")

    if not all_rows:
        print("No data collected.")
        return

    df         = pd.DataFrame(all_rows)
    df_summary = pd.DataFrame(all_summary)

    excel_path = os.path.join(output_dir, "PSU_Major_Requirements.xlsx")
    txt_path   = os.path.join(output_dir, "PSU_Major_Requirements.txt")

    print(f"\nWriting Excel: {excel_path}")
    write_excel(df, df_summary, excel_path)

    print(f"Writing TXT:   {txt_path}")
    write_txt(df, txt_path)

    print("\n" + "=" * 60)
    print("COMPLETE")
    print(f"  Programs : {df['program_name'].nunique()}")
    print(f"  Rows     : {len(df)}")
    print(f"  Excel    : PSU_Major_Requirements.xlsx")
    print(f"  TXT      : PSU_Major_Requirements.txt")
    print("=" * 60)


if __name__ == "__main__":
    main()
